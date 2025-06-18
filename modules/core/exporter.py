import sys
import os
import pandas as pd
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import shutil
import tempfile
from modules.utils import resource_file_out


def resource_path(relative_path):
    """
    Retorna o caminho absoluto para recursos, considerando o bundle do PyInstaller.
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def encontrar_proxima_linha_vazia(ws, linha_inicial=30):
    resposta_cols = ["I", "J", "K", "L", "M", "N", "O", "P", "Q", "R",
                     "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                     "AC", "AD", "AE", "AF", "AG", "AH"]
    linha = linha_inicial
    while True:
        preenchido = False
        for col in resposta_cols:
            valor = ws[f"{col}{linha}"].value
            if valor is not None and str(valor).strip() != "":
                preenchido = True
                break
        if not preenchido:
            return linha
        linha += 1

def importar_para_planilha(dados, caminho_template):
    print(f"[LOG] Função importar_para_planilha chamada para o arquivo '{caminho_template}'. Número de registros a importar: {len(dados)}")
    
    real_path = resource_path(caminho_template)

    try:
        book = load_workbook(real_path)
        print(f"[LOG] Planilha '{real_path}' carregada com sucesso.")
    except Exception as e:
        print(f"[ERRO] Falha ao carregar '{real_path}': {e}")
        return
    
    sheet_name = "GERAL"
    if sheet_name not in book.sheetnames:
        print(f"[ERRO] Aba '{sheet_name}' não encontrada na planilha.")
        return
    
    ws = book[sheet_name]
    print(f"[LOG] Usando aba '{sheet_name}'.")

    map_info_coluna = {
       "matricula": "D",
       "nome_aluno": "E",
       "escola": "G",
       "turma": "H" 
    }

    map_questao_coluna = {
        1: 'I', 2: 'J', 3: 'K', 4: 'L', 5: 'M',
        6: 'N', 7: 'O', 8: 'P', 9: 'Q', 10: 'R',
        11: 'S', 12: 'T', 13: 'U', 14: 'V', 15: 'W',
        16: 'X', 17: 'Y', 18: 'Z', 19: 'AA', 20: 'AB',
        21: 'AC', 22: 'AD', 23: 'AE', 24: 'AF', 25: 'AG',
        26: 'AH'
    }
    
    for idx, item in enumerate(dados):
        respostas = item.get("Respostas", {})
        ocr_info = item.get("OCR", {})
        
        row_inicial = encontrar_proxima_linha_vazia(ws, linha_inicial=30)

        matricula_val = ocr_info.get("matricula", "N/A")
        nome_val = ocr_info.get("nome_aluno", "N/A")
        escola_val = ocr_info.get("escola", "N/A")
        turma_val = ocr_info.get("turma", "N/A") 

        ws[f"{map_info_coluna['matricula']}{row_inicial}"] = matricula_val
        ws[f"{map_info_coluna['nome_aluno']}{row_inicial}"] = nome_val
        ws[f"{map_info_coluna['escola']}{row_inicial}"] = escola_val
        ws[f"{map_info_coluna['turma']}{row_inicial}"] = turma_val

        for questao_nome, resposta in respostas.items():
            try:
                numero = int(questao_nome.replace("Questao ", ""))
            except ValueError:
                print(f"[ERRO] Não foi possível extrair número de '{questao_nome}'")
                continue

            col_letra = map_questao_coluna.get(numero)
            if col_letra:
                ws[f"{col_letra}{row_inicial}"] = resposta
                print(f"[LOG] Inserindo {questao_nome}='{resposta}' na célula {col_letra}{row_inicial}")

        print(f"[LOG] Inserido registro {idx+1} na linha {row_inicial}")

    try:
        book.save(real_path)
        print(f"[LOG] Dados salvos na planilha '{real_path}'.")
    except Exception as e:
        print(f"[ERRO] Falha ao salvar '{real_path}': {e}")

# ============ GOOGLE SHEETS ============= #

def extrair_id_google_sheets(link_ou_id: str) -> str:
    """
    Extrai o ID de uma URL do Google Sheets ou retorna o próprio ID se for passado direto.
    """
    if "docs.google.com" in link_ou_id:
        return link_ou_id.split("/d/")[1].split("/")[0]
    return link_ou_id

def importar_para_google_sheets(dados, sheet_link_ou_id, credentials_json="credentials.json"):
    from operator import itemgetter
    sheet_id = extrair_id_google_sheets(sheet_link_ou_id)

    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    creds = Credentials.from_service_account_file(resource_file_out(credentials_json), scopes=scopes)
    client = gspread.authorize(creds)
    sh = client.open_by_key(sheet_id)

    print(f"[OK] Conectado à planilha ID: {sheet_id}")

    # ✅ Ordenar todos os dados pelo nome do aluno, insensível a maiúsculas
    dados_ordenados = sorted(dados, key=lambda x: x.get("OCR", {}).get("nome_aluno", "").lower())

    for idx, item in enumerate(dados_ordenados):
        ocr_info = item.get("OCR", {})
        respostas = item.get("Respostas", {})

        escola = ocr_info.get("escola", "SEM_ESCOLA").strip().upper()
        turma = ocr_info.get("turma", "SEM_TURMA").strip().upper()
        aba_nome = f"{escola} ({turma})"
        print(f"[INFO] Aluno {idx+1} → Aba: {aba_nome}")

        try:
            ws = sh.worksheet(aba_nome)
        except Exception:
            print(f"[WARN] Aba '{aba_nome}' não existe. Usando RESUMO GERAIS.")
            try:
                ws = sh.worksheet("RESUMO GERAIS")
            except Exception:
                print(f"[ERROR] RESUMO GERAIS também não existe. Criando fallback.")
                ws = sh.add_worksheet(title="RESUMO GERAIS", rows="1000", cols="30")

        map_colunas = {
            "matricula": "E",
            "nome_aluno": "C",
        }

        map_questoes = {
            1: "F", 2: "G", 3: "H", 4: "I", 5: "J",
            6: "K", 7: "L", 8: "M", 9: "N", 10: "O", 11: "P",
            12: "Q", 13: "R", 14: "S", 15: "T", 16: "U", 17: "V",
            18: "W", 19: "X", 20: "Y"
        }

        col_matricula = ws.col_values(5)
        linha = len(col_matricula) + 1

        ws.update(f"{map_colunas['matricula']}{linha}", [[ocr_info.get("matricula", "")]])
        ws.update(f"{map_colunas['nome_aluno']}{linha}", [[ocr_info.get("nome_aluno", "")]])

        for questao, resposta in respostas.items():
            try:
                numero = int(questao.replace("Questao ", ""))
                coluna = map_questoes.get(numero)
                if coluna:
                    ws.update(f"{coluna}{linha}", [[resposta]])
            except Exception as e:
                print(f"[WARN] Problema ao atualizar '{questao}': {e}")

        print(f"[OK] Aluno {idx+1} salvo na linha {linha} da aba '{ws.title}'")

    print("[OK] Exportação para Google Sheets concluída com sucesso")
